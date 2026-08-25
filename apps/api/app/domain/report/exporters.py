"""
Report exporters.

Exports stream. A report that fits comfortably in a preview can still be
hundreds of thousands of rows when downloaded, and materialising that in memory
would take the whole service down for everyone else -- so rows are written out
as they arrive from the server-side cursor.

Formatting is applied here rather than left to the spreadsheet: a currency
column exported as a bare float loses the meaning the report author gave it.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Iterable, Iterator

MAX_PDF_ROWS = 5_000


def _cell(value: object) -> object:
    """Native types where the format supports them, text where it does not."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ", timespec="minutes") if isinstance(
            value, datetime
        ) else value.isoformat()
    if isinstance(value, bytes):
        return "<binary>"
    return value


def to_csv(
    headers: list[str], rows: Iterable[Iterable[object]]
) -> Iterator[str]:
    """Yield CSV a chunk at a time so nothing is held in memory."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    # Excel opens UTF-8 CSV as the local codepage unless it sees a BOM, which
    # turns any non-ASCII customer name into mojibake.
    yield "﻿"

    writer.writerow(headers)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    for row in rows:
        writer.writerow([_cell(value) for value in row])
        if buffer.tell() > 32_000:
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

    if buffer.tell():
        yield buffer.getvalue()


def to_xlsx(
    headers: list[str],
    rows: Iterable[Iterable[object]],
    sheet_name: str = "Report",
    number_formats: list[str | None] | None = None,
) -> bytes:
    """
    Build an .xlsx workbook.

    Uses openpyxl's write-only mode, which keeps one row in memory at a time
    instead of the whole sheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=sheet_name[:31] or "Report")

    from openpyxl.cell import WriteOnlyCell

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5FDB")

    header_cells = []
    for title in headers:
        cell = WriteOnlyCell(sheet, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    widths = [len(str(title)) + 2 for title in headers]

    for row in rows:
        out = []
        for index, value in enumerate(row):
            cell = WriteOnlyCell(sheet, value=_cell(value))
            if number_formats and index < len(number_formats) and number_formats[index]:
                cell.number_format = number_formats[index]
            out.append(cell)
            if index < len(widths):
                widths[index] = max(widths[index], min(len(str(value or "")) + 2, 60))
        sheet.append(out)

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def to_pdf(
    headers: list[str],
    rows: list[list[object]],
    title: str = "Report",
    subtitle: str = "",
) -> bytes:
    """
    Render a landscape PDF table.

    PDF is a paginated format, so unlike CSV and Excel it cannot stream
    sensibly; the caller caps the row count before calling this.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    stream = io.BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"].clone("cell")
    cell_style.fontSize = 7.5
    cell_style.leading = 9

    header_style = cell_style.clone("head")
    header_style.textColor = colors.white
    header_style.fontName = "Helvetica-Bold"

    data = [[Paragraph(str(h), header_style) for h in headers]]
    for row in rows:
        data.append([Paragraph(str(_cell(v)), cell_style) for v in row])

    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5FDB")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F5F7FA")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    elements = [Paragraph(f"<b>{title}</b>", styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 6))
    elements.append(table)

    document.build(elements)
    return stream.getvalue()
