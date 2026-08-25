"""
Spreadsheet parsing and type inference.

Turns an uploaded CSV or Excel file into a typed table definition. Types are
inferred rather than assumed text: an amount column that arrives as text cannot
be summed, and a date column that arrives as text sorts 1 January after
31 December.

Every identifier produced here is sanitised. Column names come from a file
somebody uploaded, and they end up as SQL identifiers.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Iterator

from app.domain.schema.registry import DataType

#: How many rows to examine before deciding a column's type. Enough to be
#: confident, small enough that a large file is not read twice.
TYPE_SAMPLE_ROWS = 500

MAX_COLUMNS = 200
MAX_ROWS = 500_000

_TRUE = {"true", "yes", "y", "1", "t"}
_FALSE = {"false", "no", "n", "0", "f"}

_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y")
_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M",
    "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M:%S",
)


class UploadError(Exception):
    """A problem with the file itself. The message is shown to the user."""


@dataclass
class ParsedColumn:
    name: str          # sanitised SQL identifier
    label: str         # what the file called it
    data_type: DataType
    nullable: bool = True
    sample: list[str] = field(default_factory=list)


@dataclass
class ParsedFile:
    columns: list[ParsedColumn]
    rows: list[list]
    row_count: int
    skipped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Identifiers
# ---------------------------------------------------------------------------
_IDENTIFIER = re.compile(r"[^a-z0-9_]+")


def safe_identifier(value: str, fallback: str = "column") -> str:
    """
    Turn arbitrary text into a lowercase SQL identifier.

    Column headings arrive from a user-supplied file and become real identifiers,
    so anything outside [a-z0-9_] is replaced rather than escaped.
    """
    cleaned = _IDENTIFIER.sub("_", (value or "").strip().lower()).strip("_")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"{fallback}_{cleaned}" if cleaned else fallback
    return cleaned[:58]


def unique_identifiers(headers: list[str]) -> list[tuple[str, str]]:
    """Return (identifier, original label), guaranteeing identifiers are unique."""
    used: dict[str, int] = {}
    out: list[tuple[str, str]] = []
    for index, header in enumerate(headers):
        label = (header or "").strip() or f"Column {index + 1}"
        base = safe_identifier(label, f"column_{index + 1}")
        if base in used:
            used[base] += 1
            name = f"{base}_{used[base]}"
        else:
            used[base] = 0
            name = base
        out.append((name, label))
    return out


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------
def _as_int(value: str):
    text = value.replace(",", "").replace(" ", "")
    if not text or text in "+-":
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _as_decimal(value: str):
    text = value.replace(",", "").replace(" ", "")
    # Currency symbols and a trailing percent are common in exported sheets.
    text = re.sub(r"^[£$€₹]|%$", "", text)
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _as_date(value: str):
    for pattern in _DATE_FORMATS:
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue
    return None


def _as_datetime(value: str):
    for pattern in _DATETIME_FORMATS:
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def infer_type(values: list[str]) -> DataType:
    """
    Pick the narrowest type that fits every non-empty sample.

    Narrowest-that-fits matters: one stray "N/A" in a numeric column must make
    the whole column text rather than silently dropping that row's value.
    """
    samples = [v for v in values if v is not None and str(v).strip() != ""]
    if not samples:
        return DataType.TEXT

    lowered = [str(v).strip() for v in samples]

    if all(v.lower() in _TRUE | _FALSE for v in lowered):
        return DataType.BOOLEAN
    if all(_as_int(v) is not None for v in lowered):
        return DataType.INTEGER
    if all(_as_decimal(v) is not None for v in lowered):
        return DataType.DECIMAL
    if all(_as_date(v) is not None for v in lowered):
        return DataType.DATE
    if all(_as_datetime(v) is not None for v in lowered):
        return DataType.DATETIME
    return DataType.TEXT


def coerce(value, data_type: DataType):
    """Convert one cell, returning None when it does not fit rather than failing."""
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal, date, datetime, bool)):
        # Excel hands back real types already.
        if data_type is DataType.DATE and isinstance(value, datetime):
            return value.date()
        return value

    text = str(value).strip()
    if text == "":
        return None

    match data_type:
        case DataType.INTEGER:
            return _as_int(text)
        case DataType.DECIMAL:
            return _as_decimal(text)
        case DataType.BOOLEAN:
            lowered = text.lower()
            return True if lowered in _TRUE else False if lowered in _FALSE else None
        case DataType.DATE:
            return _as_date(text)
        case DataType.DATETIME:
            return _as_datetime(text)
    return text


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------
def _read_csv(content: bytes) -> tuple[list[str], Iterator[list]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        # Spreadsheets exported on Windows are frequently cp1252.
        text = content.decode("cp1252", errors="replace")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        headers = next(reader)
    except StopIteration as error:
        raise UploadError("That file appears to be empty.") from error
    return headers, reader


def _read_xlsx(content: bytes) -> tuple[list[str], Iterator[list]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as error:
        raise UploadError("That file could not be opened as a spreadsheet.") from error

    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(cell) if cell is not None else "" for cell in next(rows)]
    except StopIteration as error:
        raise UploadError("That spreadsheet has no rows.") from error
    return headers, (list(row) for row in rows)


def parse(filename: str, content: bytes) -> ParsedFile:
    """Read a CSV or Excel file into typed columns and coerced rows."""
    lowered = filename.lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        headers, reader = _read_xlsx(content)
    elif lowered.endswith((".csv", ".txt", ".tsv")):
        headers, reader = _read_csv(content)
    else:
        raise UploadError(
            "Only CSV and Excel files can be uploaded. Save the file as .csv or .xlsx "
            "and try again."
        )

    if len(headers) > MAX_COLUMNS:
        raise UploadError(
            f"That file has {len(headers)} columns; the limit is {MAX_COLUMNS}."
        )

    named = unique_identifiers(headers)
    width = len(named)
    if width == 0:
        raise UploadError("No column headings were found in the first row.")

    raw_rows: list[list] = []
    skipped = 0
    warnings: list[str] = []

    for row in reader:
        values = list(row)[:width]
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # blank line
        if len(values) < width:
            values += [None] * (width - len(values))
        raw_rows.append(values)
        if len(raw_rows) >= MAX_ROWS:
            warnings.append(
                f"Only the first {MAX_ROWS:,} rows were read; the rest were ignored."
            )
            break

    if not raw_rows:
        raise UploadError("That file has column headings but no data rows.")

    columns: list[ParsedColumn] = []
    for index, (name, label) in enumerate(named):
        sample = [row[index] for row in raw_rows[:TYPE_SAMPLE_ROWS]]
        data_type = infer_type([str(v) for v in sample if v is not None])
        columns.append(
            ParsedColumn(
                name=name,
                label=label,
                data_type=data_type,
                nullable=any(v is None or str(v).strip() == "" for v in sample),
                sample=[str(v) for v in sample[:3] if v is not None],
            )
        )

    rows: list[list] = []
    for raw in raw_rows:
        rows.append([coerce(raw[i], columns[i].data_type) for i in range(width)])

    return ParsedFile(
        columns=columns,
        rows=rows,
        row_count=len(rows),
        skipped_rows=skipped,
        warnings=warnings,
    )
